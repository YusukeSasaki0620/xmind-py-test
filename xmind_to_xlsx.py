#-*- coding: utf-8 -*-

import glob,os,sys
import xmind
import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.styles.alignment import Alignment

input_dir = "input_files"
output_dir = 'output_files'
END_COLUM_LABEL = PatternFill(patternType='solid', fgColor='d3d3d3')
SELL_DEFAULT_ALIGNMENT = Alignment(vertical = 'center')

def sub_topic_recursive_processing(current_sheet , sub_topic, max_column_num, current_row, current_column = 1):
  # セル書き込み
  current_cell = current_sheet.cell(current_row, current_column)
  current_cell.value = sub_topic.getTitle()
  current_cell.alignment = SELL_DEFAULT_ALIGNMENT

  sub_topics = sub_topic.getSubTopics() or []

  for sub_topic in sub_topics:
    # 深さ優先探索の再起なので、呼び出すたびにcolunm増
    current_row, max_column_num = sub_topic_recursive_processing(current_sheet, sub_topic, max_column_num, current_row, current_column + 1)

  # 終了条件：末端に来たときのみrow増加、整形用ラベルも付与
  if not bool(sub_topics):
    current_sheet.cell(current_row, current_column + 1).fill = END_COLUM_LABEL
    current_row += 1

  return current_row, max_column_num if max_column_num > current_column else current_column

if __name__ == '__main__':
  args = sys.argv
  for index, arg in enumerate(args):
    if index == 1: input_dir = arg
    elif index == 2: output_dir = arg
    else: pass

  print("input_dir:" + input_dir)
  print("output_dir:" + output_dir)
  input_file_paths = glob.glob(input_dir + '/' + "*.xmind")

  print(input_file_paths)

  for input_file_path in input_file_paths:
    file_name = os.path.splitext(os.path.basename(input_file_path))[0]
    output_file_path = output_dir + '/' + file_name + '.xlsx'

    # エクセルファイル新規作成
    wb = openpyxl.Workbook()
    # ループ処理で邪魔なので、デフォルトシートを削除
    for default_sheet in wb.worksheets:
      wb.remove(default_sheet)

    print("doing: " + input_file_path)

    # xmindファイル読み込み
    workbook = xmind.load(input_file_path)
    sheets = workbook.getSheets()
    for sheet in sheets:
      # xmindシートごとに処理
      # エクセルシート新規作成
      newSheet = wb.create_sheet(title=sheet.getTitle())
      print("start_in_sheet: " + newSheet.title)
      # 固定出力：ルートトピックは最上端
      rt = sheet.getRootTopic()
      newSheet.cell(1,1).value = rt.getTitle()

      current_row = 3 # ２行目は固定出力なので一旦飛ばす
      max_column_num = 0

      sub_topics = rt.getSubTopics() or []
      # サブトピックを再帰処理
      for sub_topic in sub_topics:
        current_row, max_column_num = sub_topic_recursive_processing(newSheet, sub_topic, max_column_num, current_row)

      # カウント確認
      # newSheet.cell(current_row,max_column_num).value = 'カウント確認'
      print("データサイズ [row: " + str(current_row) + ", col: " + str(max_column_num) + "]")

      print("セル出力 :check:")

      # ２行目ラベル出力
      for i in range(max_column_num):
        tmp_row = i+1
        newSheet.cell(2, tmp_row).value = "Level " + str(tmp_row)

      print("ラベル出力 :check:")
      # 整形処理① 末尾処理
      for row in newSheet.iter_rows(3, current_row - 1, 1, max_column_num + 1):
        find_end = False
        for cell in row:
          if find_end:
            cell.fill = END_COLUM_LABEL
          if cell.fill == END_COLUM_LABEL:
            find_end = True

      # 整形処理② 列結合
      # 先頭行はデータが有ったほうが都合が良いのでラベルから開始
      for col in newSheet.iter_cols(1, max_column_num, 2, current_row - 1):
        start_cell = end_cell= col[0]
        in_end_column_label_cell = False
        print("整形中：" + start_cell.coordinate)
        for cell in col:
          if cell == start_cell:
            # 初回スキップ
            continue

          # 通常探索
          if not in_end_column_label_cell:
            if cell.fill == END_COLUM_LABEL or bool(cell.value):
              # 末尾無効データまたは次のテキストにあたったら直前までを結合
              newSheet.merge_cells(start_cell.coordinate + ":" + end_cell.coordinate)
              start_cell = end_cell = cell
              if cell.fill == END_COLUM_LABEL:
                # 無効データに当たったなら無効セル探索に移行
                in_end_column_label_cell = True
            else :
              # 空セルの間endを移動
              end_cell = cell
          else:
            # 無効セル探索、基本無視
            start_cell = end_cell = cell
            if bool(cell.value):
              # 次のテキストが見つかったら通常探索に移行
              in_end_column_label_cell = False

        # 終点処理
        newSheet.merge_cells(start_cell.coordinate + ":" + end_cell.coordinate)


    # 上書き保存
    wb.save(output_file_path)
